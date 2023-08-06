"""
Holds the code that defines the local xlsx reader.
"""

from pathlib import Path
from typing import Callable, List, Optional, Union

import openpyxl

from tidychef.acquire.base import BaseReader
from tidychef.models.source.cell import Cell
from tidychef.models.source.table import Table
from tidychef.selection.selectable import Selectable
from tidychef.selection.xlsx.xlsx import XlsxSelectable
from tidychef.utils import fileutils

from ..base import BaseReader
from ..main import acquirer


def local(
    source: Union[str, Path],
    selectable: Selectable = XlsxSelectable,
    pre_hook: Optional[Callable] = None,
    post_hook: Optional[Callable] = None,
    tables: str = None,
    **kwargs
) -> Union[XlsxSelectable, List[XlsxSelectable]]:
    """
    Read data from a Path (or string representing a path)
    present on the same machine where tidychef is running.

    This local xlsx reader uses openpyxl:
    https://openpyxl.readthedocs.io/en/stable/index.html

    Additional kwargs passed to this function are propagated to
    the openpyxl.load_workbook() method.

    :param source: A Path object or a string representing a path
    :param selectable: A class that implements tidychef.selection.selectable.Selectable of an inheritor of. Default is XlsxSelectable
    :param pre_hook: A callable that can take source as an argument
    :param post_hook: A callable that can take the output of XlsxSelectable.parse() as an argument.
    :return: A single populated Selectable of type as specified by selectable param
    """

    assert isinstance(
        source, (str, Path)
    ), """
        The source you're passing to acquire.csv.local() needs to
        be either a Path object or a string representing such.
        """

    return acquirer(
        source,
        LocalXlsxReader(tables),
        selectable,
        pre_hook=pre_hook,
        post_hook=post_hook,
        **kwargs
    )


class LocalXlsxReader(BaseReader):
    """
    A reader to lead in a source where that source is a locally
    held xls file.
    """

    def parse(
        self,
        source: Union[str, Path],
        selectable: Selectable = XlsxSelectable,
        data_only: bool = True,
        **kwargs
    ) -> List[XlsxSelectable]:
        """
        Parse the provided source into a list of Selectables. Unless overridden the
        selectable is of type XlsxSelectable.

        Additional **kwargs are propagated to openpyxl.load_workbook()

        :param source: A Path or str representing a path indicating a local file
        :param selectable: The selectable type to be returned.
        :data_only: An openpyxl.load_workbook() option to disable acquisition of non data elements from the tabulated source (macros etc)
        :return: A list of type as specified by param selectable.
        """

        source: Path = fileutils.ensure_existing_path(source, **kwargs)

        workbook: openpyxl.Workbook = openpyxl.load_workbook(
            source, data_only=data_only, **kwargs
        )

        tidychef_selectables = []
        worksheet_names = workbook.get_sheet_names()
        for worksheet_name in worksheet_names:

            worksheet = workbook.get_sheet_by_name(worksheet_name)

            table = Table()
            for y, row in enumerate(worksheet.iter_rows()):
                for x, cell in enumerate(row):
                    table.add_cell(
                        Cell(x=x, y=y, value=str(cell.value) if cell.value else "")
                    )

            tidychef_selectables.append(
                selectable(table, source=source, name=worksheet_name)
            )
        return tidychef_selectables
