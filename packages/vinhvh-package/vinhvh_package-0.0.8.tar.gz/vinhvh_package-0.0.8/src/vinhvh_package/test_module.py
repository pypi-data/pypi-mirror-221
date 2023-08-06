
import PySimpleGUI as sg
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import time
import os
from pathlib import Path
import logging
import sys


def select_file(names, mode='single'):
    # Set the PySimpleGUI theme
    sg.theme('Dark Blue 3')

    # Create the layout for the file selector window
    layout = [[sg.Text('Choose files ')]]

    # Add input fields and file browse buttons for each name in names
    if mode == 'single':
        for name in names:
            layout.append([sg.Text(name + ": ", size=(15, 1),),
                           sg.InputText(), sg.FileBrowse()])
    # Add input fields and files browse buttons for each name in names
    elif mode == 'multi':
        for name in names:
            layout.append([sg.Text(name + ": ", size=(15, 1),),
                           sg.InputText(key=f'{name}'), sg.FilesBrowse()])

    # Add the Submit and Cancel buttons to the layout
    layout.append([sg.Submit(), sg.Cancel()])

    # Create the file selector window
    window = sg.Window('File Selector', layout)

    # Wait for the user to interact with the window
    event, values = window.read()

    # Close the window
    window.close()

    # Process the user's selection
    if event == 'Submit':
        # If in single mode, get the file paths from the input fields
        if mode == 'single':
            file_paths = [values[i] for i in range(len(names))]
        # If in multi mode, get the file paths as a list of lists from the input fields
        elif mode == 'multi':
            file_paths = [values[f'{name}'].split(';') for name in names]

        # Return the list of file paths
        return file_paths
    # If the user clicked the Cancel button, print a message and exit the program
    elif event == 'Cancel':
        print('File selection cancelled.')
        sys.exit()
    # If the user closed the window, print a message and exit the program
    elif event == sg.WIN_CLOSED:
        print('File selection closed.')
        sys.exit()
    # If the user didn't select any files, return None
    else:
        return None


def select_file_2(file: list = None, value: list = None, option: dict = None, checkbox: list = None, mode: str = 'single', title: str = 'File Selector'):
    """
    This function create a interface for user with different option input

    Parameters
    ----------
    file: (list, optional)
        File to browse
    value: (list, optional)
        Value for input
    Option: (dict, optional)
        Combobox for select
    Checkbox: (list, optional)
        Option for select
    Tilte: (str, optional)
        Name of window
    """
    # Set the PySimpleGUI theme
    sg.theme('Dark Blue 3')

    # Create the layout for the file selector window
    layout = [[sg.Text('Select values ')]]
    file_paths = []
    len_path = 0
    if file is not None:
        if mode == 'single':
            for name in file:
                layout.append([sg.Text(name + ": ", size=(15, 1),),
                               sg.InputText(), sg.FileBrowse()])
        elif mode == 'multi':
            for name in file:
                layout.append([sg.Text(name + ": ", size=(15, 1),),
                               sg.InputText(key=f'{name}'), sg.FilesBrowse()])

    if value is not None:
        for name in value:
            layout.append([sg.Text(name + ": ", size=(15, 1),),
                           sg.InputText()])

    if option is not None:
        for name, value in option.items():
            layout.append([sg.Text(name + ': ', size=(15, 1)),
                          sg.Combo(value, default_value=value[0], size=(43, 1))],)

    if checkbox is not None:
        for name in checkbox:
            layout.append([sg.Checkbox(name + ':', default=False)])

    # Add the Submit and Cancel buttons to the layout
    layout.append([sg.Submit(), sg.Cancel()])

    # Create the file selector window
    window = sg.Window(title, layout)

    # Wait for the user to interact with the window
    event, values = window.read()

    # Close the window
    window.close()

    # Process the user's selection
    if event == 'Submit':

        if file is not None:
            # If in single mode, get the file paths from the input fields
            if mode == 'single':
                path = [values[i] for i in range(len(file))]
                len_path += len(file)

            # If in multi mode, get the file paths as a list of lists from the input fields
            elif mode == 'multi':
                path = [values[f'{name}'].split(';') for name in file]
            file_paths += path

        if value is not None:
            path = [values[i] for i in range(len_path, len_path + len(value))]
            file_paths += path
            len_path += len(value)

        if option is not None:
            path = [values[i] for i in range(len_path, len_path + len(option))]
            file_paths += path
            len_path += len(option)

        if checkbox is not None:
            path = [values[i]
                    for i in range(len_path, len_path + len(checkbox))]
            file_paths += path
            len_path += len(checkbox)

            # Return the list of file paths
        return file_paths

    # If the user clicked the Cancel button, print a message and exit the program
    elif event == 'Cancel':
        print('File selection cancelled.')
        sys.exit()
    # If the user closed the window, print a message and exit the program
    elif event == sg.WIN_CLOSED:
        print('File selection closed.')
        sys.exit()
    # If the user didn't select any files, return None
    else:
        return None


def select_input_file(names1, names2, mode='single'):
    # Set the PySimpleGUI theme
    sg.theme('Dark Blue 3')

    # Create the layout for the file selector window
    layout = [[sg.Text('Choose files ')]]

    # Add input fields and file browse buttons for each name in names1
    if mode == 'single':
        for name in names1:
            layout.append([sg.Text(name + ": ", size=(15, 1),),
                           sg.InputText(), sg.FileBrowse()])
    # Add input fields and files browse buttons for each name in names1
    elif mode == 'multi':
        for name in names1:
            layout.append([sg.Text(name + ": ", size=(15, 1),),
                           sg.InputText(key=f'{name}'), sg.FilesBrowse()])

    # Add input fields for each name in names2
    for name in names2:
        layout.append([sg.Text(name + ": ", size=(15, 1),),
                       sg.InputText()])

    # Add the Submit and Cancel buttons to the layout
    layout.append([sg.Submit(), sg.Cancel()])

    # Create the file selector window
    window = sg.Window('File Selector', layout)

    # Wait for the user to interact with the window
    event, values = window.read()

    # Close the window
    window.close()

    # Process the user's selection
    if event == 'Submit':
        # If in single mode, get the file paths and input values from the input fields
        if mode == 'single':
            file_paths = [values[i] for i in range(len(names1))]
            input = [values[i]
                     for i in range(len(names1), len(names1) + len(names2))]

        # If in multi mode, get the file paths as a list of lists and the input values from the input fields
        elif mode == 'multi':
            file_paths = [values[f'{name}'].split(';') for name in names1]
            input = [values[i] for i in range(len(names2))]

        # Combine the file paths and input values into a single list
        file_paths = file_paths + input

        # Return the list of file paths and input values
        return file_paths
    # If the user clicked the Cancel button, print a message and exit the program
    elif event == 'Cancel':
        print('File selection cancelled.')
        sys.exit()
    # If the user closed the window, print a message and exit the program
    elif event == sg.WIN_CLOSED:
        print('File selection closed.')
        sys.exit()
    # If the user didn't select any files, return None
    else:
        return None


def read_file(filepath):
    # Determine file type based on extension
    if filepath.lower().endswith(('.txt', '.csv', '.xlsx', '.xls')):
        if filepath.lower().endswith('.txt'):
            df = pd.read_csv(filepath, sep='	', low_memory=False)
        elif filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath, low_memory=False)
        elif filepath.lower().endswith('.xlsx'):
            df = pd.read_excel(filepath, engine='openpyxl')
        elif filepath.lower().endswith('.xls'):
            df = pd.read_excel(filepath, engine='xlrd')
        return df
    else:
        raise ValueError("Unsupported file format")


def read_logfile(filepath):
    if filepath.lower().endswith('.log'):
        with open(filepath, 'r') as f:
            lines = f.readlines()

        # create a dataframe from the list of strings
        df_log = pd.DataFrame(lines, columns=['log'])
        return df_log
    else:
        raise ValueError("Unsupported file format")


def save_file(df_result, file_name, mode="Result"):
    # Generate a timestamp for the current date and time
    date_time = datetime.now().strftime("%d%m%y_%H%M%S")

    # Append the timestamp to the filename to make it unique
    file_name = f'{file_name} {date_time}.xlsx'

    if mode == "Result":
        # Create the result directory if it doesn't exist
        Path("Result").mkdir(exist_ok=True)

        # Change the current working directory to the result directory
        os.chdir("Result")

    # Save the DataFrame to an Excel file without the index
    df_result.to_excel(file_name, index=False)

    # Get the full file path of the saved file
    file_path = os.path.join(os.getcwd(), file_name)

    if mode == "Result":
        # Change the current working directory back to the main source directory
        os.chdir("..")

    # Return the full file path of the saved file
    return file_path


def format_excel(file_name):
    # Load the Excel spreadsheet
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Set the font and alignment styles for all cells
    font = Font(name='Calibri', size=11)
    alignment = Alignment(horizontal='center',
                          vertical='center')

    # Add borders to all cells
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Set the background color and font and alignment styles for the first row
    fill = PatternFill(start_color='FFC000',
                       end_color='FFC000', fill_type='solid')
    font_first_row = Font(name='Calibri', size=11, bold=True)
    alignment_first_row = Alignment(
        horizontal='center', vertical='center', wrap_text=True)

    # Set all columns to the same width as the original file
    for i, column in enumerate(ws.columns):
        ws.column_dimensions[column[0].column_letter].width = 15

    # Format all cells in the worksheet
    for row in ws.rows:
        for cell in row:
            if cell.row == 1:
                # Format the first row cells
                cell.font = font_first_row
                cell.alignment = alignment_first_row
                cell.border = border
                cell.fill = fill
            else:
                # Format all other cells
                cell.font = font
                cell.alignment = alignment
                cell.border = border

    # Save the modified Excel spreadsheet
    wb.save(file_name)
    print('Finish create excel file!')


def check_license(start_date_str, duration_days):
    # Convert the start date string to a datetime object
    start_date = datetime.fromisoformat(start_date_str)

    # Calculate the end date based on the start date and duration in days
    end_date = start_date + timedelta(days=duration_days)

    # Get the current date and time
    now_date = datetime.now()

    # Calculate the time difference between the end date and the current date in seconds
    time_compare = end_date.timestamp() - now_date.timestamp()

    # Calculate the number of whole days remaining in the license period
    days_remaining = int(time_compare / 86400)

    # Determine the status of the license based on the number of days remaining
    if days_remaining < 0:
        print('Your program license has been expired, please contact to provider for more information!')
        return False
    elif days_remaining == 0 or days_remaining == 1:
        print(f'License: {days_remaining} day')
        return True
    elif days_remaining <= 30 and days_remaining > 1:
        print('Lic = True')
        print(f'License: {days_remaining} days')
        return True
    else:
        print('Lic = True')
        return True


def popup_finish(program_running_time, result_path):
    # Create the message to display in the popup window
    message = '        Finish!' + '            ' + '\n' +\
        '        Program running time: ' + program_running_time + '            ' + '\n' +\
        '        Result: ' + result_path + '\n' +\
        '        Copyright: VinhVH      '

    # Display the popup window with the message and a title
    sg.popup_auto_close(
        message, title='Program running result', auto_close_duration=20)


def popup_error():
    # Create the result directory if it doesn't exist
    Path("Logfile").mkdir(exist_ok=True)
    # Change the current working directory to the Logfile directory
    os.chdir("Logfile")
    date_time = datetime.now().strftime("%d%m%y_%H%M%S")
    logfile = f'Error logfile {date_time}.log'

    # Configure the logging module
    logging.basicConfig(filename=logfile, level=logging.ERROR)

    try:
        # Your code that may raise an exception
        raise Exception("Something went wrong")
    except Exception as e:
        # Log the exception to the file
        logging.exception("An exception was thrown: %s", str(e))
        # Print the exception to the terminal
        print("An exception was thrown: ", str(e))
        # Print the exception traceback to the terminal
        exc_type, exc_value, exc_traceback = sys.exc_info()
        traceback_details = {
            'filename': exc_traceback.tb_frame.f_code.co_filename,
            'lineno': exc_traceback.tb_lineno,
            'name': exc_traceback.tb_frame.f_code.co_name,
            'type': exc_type.__name__,
            'message': str(exc_value)
        }
        for key, value in traceback_details.items():
            print(f"{key}: {value}")

    message = f'        Error \n        Bye bye!'
    sg.popup_auto_close(message, title='Notice',
                        auto_close_duration=20, background_color='red')
    time.sleep(10)


def start():
    st_time = time.time()
    start_time = datetime.now()
    print('Program is running...')
    print('Start running time:', start_time.strftime("%d/%m/%Y, %H:%M:%S"))
    return st_time


def end(st_time, result_path):
    end_time = time.time()
    finish_time = datetime.now()
    print('End running time:', finish_time.strftime("%d/%m/%Y, %H:%M:%S"))
    running_time = end_time - st_time
    program_running_time = f"{int(running_time//60)}'" + \
        f'{int(running_time%60)}s'
    print('Program running time:', program_running_time)
    print('Mission completed!')
    message = f'        Finish!\n        Program running time: {program_running_time}\n        Result path: {result_path}\n        Copyright: VinhVH'
    sg.popup_auto_close(
        message, title='Running result', auto_close_duration=20)


def the_most_common_value(df):
    common_values = {}
    for col in df.columns:
        mode = df[col].mode()
        if not mode.empty:
            common_values[col] = mode.iloc[0]
        else:
            common_values[col] = 'Empty'
    return common_values


def type_site(x):
    site_types = {
        'IBC': ['7', '8', '9', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'I4BA', 'I4BB', 'I4BC', 'I4BD', 'I4BE', 'I4BF', 'I4BJ', 'I4BK', 'I4CA', 'I4CB', 'I4CC', 'I4CD', 'I4CE', 'I4CF', 'I4CJ', 'I4CK'],
        'Macro': ['1', '2', '3', '4', '5', '6', 'A', 'B', 'D', 'E', 'F', 'G', 'H', 'I', 'M4BA', 'M4BB', 'M4BC', 'M4BD', 'M4BE', 'M4BF', 'M4BG', 'M4BH', 'M4BI', 'M4CA', 'M4CB', 'M4CC', 'M4CD', 'M4CE', 'U4BA', 'U4BB', 'U4BC', 'U4CA', 'U4CB', 'U4CC', 'A_LTE', 'B_LTE', 'C_LTE', 'D_LTE', 'E_LTE', 'F_LTE'],
        'CRAN - indoor': ['J4BA', 'J4BB', 'J4BC', 'J4CA', 'J4CB', 'J4CC'],
        'CRAN - outdoor': ['C4BA', 'C4BB', 'C4BC', 'C4CA', 'C4CB', 'C4CC'],
        'Smallcell': ['S']
    }
    for site_type, codes in site_types.items():
        if x in codes:
            return site_type
    return 'Other'


def dvt(x):
    site_types = {
        'DVTD': ['H01', 'H02', 'H03', 'H04', 'H05', 'H07', 'H09', 'H10', 'H11', 'HBT', 'HCG', 'HNB', 'HPN', 'HTD'],
        'DVTT': ['H06', 'H08', 'H12', 'HBC', 'HBI', 'HCC', 'HGV', 'HHM', 'HTB', 'HTP']
    }
    for site_type, codes in site_types.items():
        if x in codes:
            return site_type
    return 'Other'
